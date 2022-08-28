# philencegold - 2022

from PyPDF2 import PdfFileReader
import re
import unidecode
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Color, PatternFill, Font, Border

class EvalAuto:

    def __init__(self, reponses_eleves='', reponses_corrige=''):
        self.note = 0
        self.debug = True
        if reponses_eleves != '':
            self.reponses_eleve = self.reponses(reponses_eleves)
        if reponses_corrige != '':
            self.reponses_corrige = self.reponses(reponses_corrige)
        self.list_eval = []
        self.headers = []
        if self.debug:
            if reponses_eleves != '':
                print(self.reponses_eleve)
            if reponses_corrige != '':
                print(self.reponses_corrige)

    def getNote(self):
        return self.note


    def reponses(self, pdf_file_name) :
        f = PdfFileReader(pdf_file_name)
        fields = f.getFields()
        fdfinfo = dict((k, v.get('/V', '')) for k, v in fields.items())
        return fdfinfo

    def keepNumbers(self, s):
        r = re.findall(r'\d+', s)
        return "".join(r)

    def removeNumbers(self, s):
        return "".join((x for x in s if not x.isdigit()))

    def normalizeChain(self, s, removeSpaces = True):
        r = s.lower()
        if removeSpaces:
            r=r.replace(" ","")
        r=r.replace("-","")
        r=r.replace("'"," ")
        r = unidecode.unidecode(r)
        return "".join(r)



    def normalizeList(self,liste) :
        res=[]
        for mots in liste:
            m=self.normalizeChain(mots)
            m = re.sub(r'[^a-zA-Z]', '', m)
            res.append(m)
        return res

## Creer une liste normalisee des mots d'une phrase

    # def list_mots(self, s):
    #     l = s.split()
    #     res=[]
    #     for mots in l:
    #         m=self.normalizeChain(mots)
    #         m = re.sub(r'[^a-zA-Z]', '', m)
    #         if m!='':
    #             res.append(m)
    #     return res

## Nombre et listes des éléments communs à deux listes

    # def motsCommuns(self, list1, list2):
    #     a = set(list1)
    #     b = set(list2)
    #     c = a.intersection(b)
    #     return [len(c),c]

    def motsPresents(self,rep,liste) :
        chaine=self.normalizeChain(rep, False)
        reference=self.normalizeList(liste)
        cpt=0
        mots=[]
        for m in reference :
            pattern='.*%s'%m
            if re.match(pattern,chaine) :
                cpt+=1
                mots.append(m)
        return [cpt,mots]
            

    def distance(self, rep_a,rep_b,type):
        if type == "nombres" :
            a = self.keepNumbers(rep_a)
            b = self.keepNumbers(rep_b)
            if (a != '') :
                d = [abs(int(a)-int(b))]
            else :
                d = [-1]
        if type == "chaine" :
            a = self.normalizeChain(rep_a)
            b = self.normalizeChain(rep_b)
            if (a != '') and a == b :
                d = [0]
            else :
                d = [-1]
        if type == "mots" :
            if (rep_a != '') :
                d = rep_a
            else :
                d = [-1]

        if type == "mixte":
            a1 = self.keepNumbers(rep_a)
            b1 = self.keepNumbers(rep_b)
            if (a1 != '') :
                d1 = abs(int(a1)-int(b1))
            else:
                d1 = -1
            a2 = self.normalizeChain(self.removeNumbers(rep_a))
            b2 = self.normalizeChain(self.removeNumbers(rep_b))
            if (a2 != '') and  a2 == b2 :
                d2 = 0
            else :
                d2 = -1
            d=[d1,d2]   
        if type == "texte" :
            d=[rep_a]
        
        return d


    def evaluation(self, n, type, borne):
        rep = 'champ'+str(n)
        rep_a = self.reponses_eleve[rep]
        rep_b = self.reponses_corrige[rep]
               
        if n <= 3:
            self.list_eval.append(rep_a)
        else :
            if type ==  "texte" :
                self.headers.append("Q%s"%(n-3))
                self.list_eval.append( {'note': '', 'comm': rep_a}) 
            elif type == "mots" :
                d = self.distance(rep_a,rep_b,type)
                self.headers.append("Q%s"%(n-3))
                if  d != -1 :
                    ref=self.normalizeList(borne)
                    mots=self.motsPresents(d,ref)
                    self.list_eval.append({'note': mots[0], 'comm': rep_a})
                    print("%s : OK %s"%(rep,mots[1]))
                else :
                    self.list_eval.append({'note': 0, 'comm': rep_a})
                    print("%s : %s"%(rep, rep_a))
            else :
                d = self.distance(rep_a,rep_b,type)
                for i in range(len(d)) :
                    self.headers.append("Q%s"%(n-3))
                    if  d[i] <= borne[i] and d[i] != -1 :
                        self.list_eval.append({'note': 1, 'comm': rep_a})
                        print("%s : OK"%rep)
                    else :
                        self.list_eval.append({'note': 0, 'comm': rep_a})
                        print("%s : %s"%(rep, rep_a))




    def getNotes(self):
        return self.list_eval

    def getHeaders(self):
        return  self.headers

    @staticmethod
    def writeToExcel(res_classe):
        wb = Workbook()
        f= 'notes.xlsx'
        ws1 = wb.active
        ws1.title = "Notes"
        for r in range(0,len(res_classe)):
            for c in range(0,len(res_classe[r])):     
                if r == 0:
                    column_letter = get_column_letter(c+4)
                    ws1["%s%s"%(column_letter, r+1)].value = res_classe[r][c]
                    ws1["%s%s"%(column_letter, r+1)].font = Font(bold=True)
                    ws1["%s%s"%(column_letter, r+1)].alignment = Alignment(horizontal='center',vertical='center')
                else:
                    column_letter = get_column_letter(c+1)
                    if c < 3:
                        ws1["%s%s"%(column_letter, r+2)].value = res_classe[r][c]
                        ws1["%s%s"%(column_letter, r+2)].font = Font(bold=True)
                        ws1["%s%s"%(column_letter, r+2)].fill = PatternFill("solid", start_color="c3e7eb")
                    else:
                        ws1["%s%s"%(column_letter, r+2)].value = res_classe[r][c]['note']
                        comment = Comment(res_classe[r][c]['comm'], "Author") 
                        ws1["%s%s"%(column_letter, r+2)].comment = comment
                        ws1["%s%s"%(column_letter, r+2)].number_format = '0'
                        ws1["%s%s"%(column_letter, r+2)].alignment = Alignment(horizontal='center',vertical='center')

            # Note de l'élève
            if r != 0:
                ws1["%s%s"%(get_column_letter(len(res_classe[r])+2), r+2)] = "=SUMPRODUCT(D2:%s2,D%s:%s%s)"%(get_column_letter(len(res_classe[r])), r+3, get_column_letter(len(res_classe[r])), r+3)
                ws1["%s%s"%(get_column_letter(len(res_classe[r])+2), r+2)].alignment = Alignment(horizontal='center',vertical='center')
                ws1["%s%s"%(get_column_letter(len(res_classe[r])+2), r+2)].font = Font(bold=True,color = 'FF0000')

        # Formattage conditionnel sur les notes : fond vert si notation manuelle
        green_fill = PatternFill(bgColor="cdd9c7")
        plage_result='D4:%s%s'%(get_column_letter(len(res_classe[1])),(len(res_classe)+2))
        ws1.conditional_formatting.add(plage_result,  FormulaRule(formula=['ISBLANK(D4)'], stopIfTrue=True, fill=green_fill))


        # Formattage conditionnel sur les notes : fond rouge si réponse fausse
        red_fill = PatternFill(bgColor="fad7d7")
        plage_result='D4:%s%s'%(get_column_letter(len(res_classe[1])),(len(res_classe)+2))
        ws1.conditional_formatting.add(plage_result, CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=red_fill))


        # insertion d'une ligne pour le barème : par défaut 1 par élément évalué
        ws1.insert_rows(2)
        ws1['C2'].value= 'barème'
        ws1['C2'].font=Font(italic=True,color = '0000FF')
        ws1['C2'].alignment = Alignment(horizontal='center',vertical='center')
        for c in range(0,(len(res_classe[r])-3)):
            column_letter = get_column_letter(c+4)
            ws1["%s%s"%(column_letter, 2)].value = 1
            ws1["%s%s"%(column_letter, 2)].font=Font(italic=True,color = '0000FF')
            ws1["%s%s"%(column_letter, 2)].alignment = Alignment(horizontal='center',vertical='center')
        ws1["%s2"%(get_column_letter(len(res_classe[1])+2))].value= 'Note'
        ws1["%s2"%(get_column_letter(len(res_classe[1])+2))].font=Font(bold=True,color = 'FF0000')
        ws1["%s2"%(get_column_letter(len(res_classe[1])+2))].alignment = Alignment(horizontal='center',vertical='center')

        wb.save(filename = f)


